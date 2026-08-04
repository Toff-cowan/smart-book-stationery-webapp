"""Excel backup/export and import for inventory products."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.extensions.db import db
from app.models import Product

EXPORT_HEADERS = [
    "id",
    "name",
    "department",
    "quantity",
    "price",
    "description",
    "author",
    "publisher",
    "vendor",
    "isbn",
    "school",
    "grades",
    "image_url",
    "is_active",
]

_HEADER_ALIASES = {
    "id": "id",
    "name": "name",
    "title": "name",
    "department": "department",
    "dept": "department",
    "quantity": "quantity",
    "qty": "quantity",
    "stock": "quantity",
    "price": "price",
    "description": "description",
    "desc": "description",
    "author": "author",
    "publisher": "publisher",
    "vendor": "vendor",
    "isbn": "isbn",
    "upc": "isbn",
    "school": "school",
    "grades": "grades",
    "grade": "grades",
    "image_url": "image_url",
    "image": "image_url",
    "is_active": "is_active",
    "active": "is_active",
}


def build_inventory_workbook(products: list[Product]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(EXPORT_HEADERS)
    for product in products:
        grades = product.get_grades()
        ws.append(
            [
                product.id,
                product.name,
                product.department,
                product.stock,
                float(product.price) if product.price is not None else 0,
                product.description or "",
                product.author or "",
                product.publisher or "",
                product.vendor or "",
                product.isbn or "",
                product.school or "",
                "; ".join(grades),
                product.image_url or "",
                "yes" if product.is_active else "no",
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().casefold()
    if text in {"1", "true", "yes", "y", "active"}:
        return True
    if text in {"0", "false", "no", "n", "inactive"}:
        return False
    return default


def _parse_grades(value: Any) -> list[str]:
    text = _cell_str(value)
    if not text:
        return []
    parts = [p.strip() for p in text.replace("|", ";").split(";")]
    return [p for p in parts if p]


def _parse_price(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace("$", "").replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(row):
        key = _HEADER_ALIASES.get(str(raw or "").strip().casefold())
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _row_dict(headers: dict[str, int], row: tuple[Any, ...]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, idx in headers.items():
        out[key] = row[idx] if idx < len(row) else None
    return out


def import_inventory_workbook(file_bytes: bytes) -> dict[str, Any]:
    """
    Upsert products from an Excel backup.

    Match order: id → isbn → create new.
    Returns counts: created, updated, skipped, errors (list of messages).
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": ["Spreadsheet is empty"],
        }

    headers = _header_map(tuple(header_row))
    if "name" not in headers:
        return {
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": ["Missing required column: name"],
        }

    created = updated = skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        data = _row_dict(headers, tuple(row))
        name = _cell_str(data.get("name"))
        if not name:
            skipped += 1
            errors.append(f"Row {row_num}: missing name")
            continue

        department = (_cell_str(data.get("department")) or "textbooks").casefold()
        if department not in Product.DEPARTMENTS:
            errors.append(
                f"Row {row_num}: invalid department '{department}' — skipped"
            )
            skipped += 1
            continue

        price = _parse_price(data.get("price"))
        quantity = _parse_int(data.get("quantity"))
        product_id = _parse_int(data.get("id"))
        isbn = _cell_str(data.get("isbn"))

        product: Product | None = None
        if product_id:
            product = db.session.get(Product, product_id)
        if product is None and isbn:
            product = Product.query.filter(
                Product.isbn.isnot(None),
                Product.isbn == isbn,
            ).first()

        try:
            if product is None:
                if price is None:
                    price = Decimal("0.00")
                if quantity is None:
                    quantity = 0
                product = Product(
                    name=name,
                    department=department,
                    price=price,
                    stock=max(quantity, 0),
                    description=_cell_str(data.get("description")),
                    author=_cell_str(data.get("author")),
                    publisher=_cell_str(data.get("publisher")),
                    vendor=_cell_str(data.get("vendor")),
                    isbn=isbn,
                    school=_cell_str(data.get("school")),
                    image_url=_cell_str(data.get("image_url")),
                    is_active=_parse_bool(data.get("is_active"), True),
                )
                db.session.add(product)
                db.session.flush()
                if "grades" in data:
                    product.set_grades(_parse_grades(data.get("grades")))
                created += 1
            else:
                product.name = name
                product.department = department
                if price is not None:
                    product.price = price
                if quantity is not None:
                    product.stock = max(quantity, 0)
                if "description" in data:
                    product.description = _cell_str(data.get("description"))
                if "author" in data:
                    product.author = _cell_str(data.get("author"))
                if "publisher" in data:
                    product.publisher = _cell_str(data.get("publisher"))
                if "vendor" in data:
                    product.vendor = _cell_str(data.get("vendor"))
                if "isbn" in data:
                    product.isbn = isbn
                if "school" in data:
                    product.school = _cell_str(data.get("school"))
                if "image_url" in data:
                    product.image_url = _cell_str(data.get("image_url"))
                if "is_active" in data:
                    product.is_active = _parse_bool(data.get("is_active"), product.is_active)
                if "grades" in data:
                    product.set_grades(_parse_grades(data.get("grades")))
                updated += 1
        except Exception as exc:  # noqa: BLE001 — collect row errors for admin
            skipped += 1
            errors.append(f"Row {row_num}: {exc}")

    db.session.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:40],
    }
